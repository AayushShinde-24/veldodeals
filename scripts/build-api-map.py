"""Generate Veldo-API-Requirements.xlsx — a table the founder can fill in with keys."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- palette ----
HDR = PatternFill("solid", fgColor="1F2937")      # slate-800
CORE = PatternFill("solid", fgColor="FDE68A")     # amber
SALES = PatternFill("solid", fgColor="BFDBFE")    # blue
FUND = PatternFill("solid", fgColor="C7D2FE")     # indigo
MKTG = PatternFill("solid", fgColor="FBCFE8")     # pink
BILL = PatternFill("solid", fgColor="A7F3D0")     # green
INFRA = PatternFill("solid", fgColor="E5E7EB")    # gray
white = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

# ======================================================================
# Sheet 1 — API Requirements
# ======================================================================
ws = wb.active
ws.title = "API Requirements"

cols = ["Priority Area", "Feature / Capability", "API / Service",
        "Env Var(s)", "Required?", "What it unlocks",
        "Behavior WITHOUT the key", "Wired? (real API call in code)",
        "YOUR KEY / STATUS (fill in)"]
ws.append(cols)
for c in range(1, len(cols) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HDR; cell.font = white; cell.alignment = wrap; cell.border = border

fill_by_area = {
    "CORE (all 3)": CORE, "SALES": SALES, "FUNDRAISING": FUND,
    "MARKETING": MKTG, "BILLING": BILL, "INFRA": INFRA,
}

rows = [
    # ---- CORE ----
    ["CORE (all 3)", "Database / auth / persistence (leave demo mode)", "Supabase",
     "NEXT_PUBLIC_SUPABASE_URL; NEXT_PUBLIC_SUPABASE_ANON_KEY; SUPABASE_SERVICE_ROLE_KEY",
     "REQUIRED", "Real users, saved leads/campaigns/emails/deals. Everything real depends on this.",
     "App runs in DEMO mode = fake sample data only. Nothing is saved.",
     "YES", ""],
    ["CORE (all 3)", "All AI: email writer, ad copy, research, reply classify, scoring", "Anthropic (primary)",
     "ANTHROPIC_API_KEY; ANTHROPIC_MODEL_ADVANCED; ANTHROPIC_MODEL_PREMIUM",
     "REQUIRED", "Every agent's real generation. Hyper-personalization uses the premium (deep) tier.",
     "Templated/generic text or 'No AI provider configured' errors on AI routes.",
     "YES", ""],
    ["CORE (all 3)", "AI failover provider", "OpenAI",
     "OPENAI_API_KEY; OPENAI_MODEL", "Optional",
     "Automatic fallback if Anthropic fails. Either provider satisfies the AI requirement.",
     "No failover; relies solely on Anthropic.", "YES", ""],

    # ---- SALES ----
    ["SALES", "Lead Finder — source + enrich B2B leads", "Apollo.io",
     "APOLLO_API_KEY", "REQUIRED for auto lead-gen",
     "Pull real verified contacts by title/industry/size into campaigns.",
     "Returns 0 leads; you must upload CSV manually.", "YES", ""],
    ["SALES", "Company research + buying signals (feeds personalization)", "Tavily",
     "TAVILY_API_KEY", "Recommended",
     "Real-time web research per account -> the 'why now' behind hyper-personalized emails.",
     "Research/signals skipped; emails less specific.", "YES", ""],
    ["SALES", "Email verification before send", "ZeroBounce",
     "ZEROBOUNCE_API_KEY", "Recommended",
     "Validates each address; protects deliverability/sender reputation.",
     "Every email optimistically marked 'valid' (no real check).", "YES", ""],
    ["SALES", "Send email + book meetings (Gmail/Calendar)", "Google OAuth",
     "GOOGLE_CLIENT_ID; GOOGLE_CLIENT_SECRET; GOOGLE_REDIRECT_URI", "REQUIRED to send",
     "Connect a real mailbox, send/track from it, auto-create calendar events.",
     "No real outbound email; sending is simulated.", "YES", ""],
    ["SALES", "Transactional / non-Gmail send fallback", "Resend",
     "RESEND_API_KEY", "Optional",
     "Fallback sending channel when a Gmail mailbox isn't connected.",
     "No fallback channel.", "YES", ""],
    ["SALES", "AI cold calling", "Vapi / Bland / Retell",
     "VOICE_PROVIDER_API_KEY; VOICE_PROVIDER", "Optional",
     "Places real AI voice calls to leads (compliance-gated).",
     "Mock call handle only — no one is dialed.", "YES", ""],
    ["SALES", "Proposal e-signature (deal close)", "E-sign provider",
     "ESIGN_PROVIDER_API_KEY; ESIGN_PROVIDER; ESIGN_API_URL", "Optional",
     "Send proposals for real e-signature.",
     "Proposal generated but not sent for signature.", "PARTIAL", ""],

    # ---- FUNDRAISING ----
    ["FUNDRAISING", "Pitch drafting, investor outreach, raise-readiness", "Anthropic/OpenAI (shared)",
     "(uses ANTHROPIC_API_KEY / OPENAI_API_KEY above)", "REQUIRED",
     "Real AI-written pitches + personalized investor outreach + readiness scoring.",
     "Generic/templated output.", "YES", ""],
    ["FUNDRAISING", "Live investor database (source matched investors)", "Investor DB (e.g. Harmonic/Crunchbase)",
     "INVESTOR_DB_API_KEY", "Optional",
     "Would pull a live investor universe to match against your raise.",
     "Uses a small built-in CURATED investor list (matching still works).",
     "NOT WIRED — flag only; no external call yet", ""],
    ["FUNDRAISING", "Investor outreach email send", "Google OAuth / Resend (shared)",
     "(same as SALES sending keys)", "REQUIRED to send",
     "Send investor emails from a real mailbox.",
     "Simulated send.", "YES", ""],

    # ---- MARKETING ----
    ["MARKETING", "Ad copy generation (per channel)", "Anthropic/OpenAI (shared)",
     "(uses ANTHROPIC_API_KEY / OPENAI_API_KEY above)", "REQUIRED",
     "Real, channel-tailored ad headlines + body copy.",
     "Templated placeholder copy.", "YES", ""],
    ["MARKETING", "Ad creative media (image/video)", "fal.ai (Seedance 2.0 / Flux)",
     "FAL_KEY; FAL_AD_VIDEO_MODEL; FAL_AD_IMAGE_MODEL", "Optional",
     "Generate the actual image/video creative.",
     "Returns a text creative CONCEPT only; media URL stays empty.",
     "NOT WIRED — key detected but fal job not called yet", ""],
    ["MARKETING", "Publish ads to Meta (FB/IG)", "Meta Ads API",
     "META_ADS_ACCESS_TOKEN; META_AD_ACCOUNT_ID", "Optional",
     "Push campaigns live to Meta.",
     "Marked 'published' in DB only — no real platform post.",
     "NOT WIRED — publish is DB-stub", ""],
    ["MARKETING", "Publish ads to Google", "Google Ads API",
     "GOOGLE_ADS_DEVELOPER_TOKEN; GOOGLE_ADS_CUSTOMER_ID", "Optional",
     "Push campaigns live to Google.",
     "DB-stub only.", "NOT WIRED — publish is DB-stub", ""],
    ["MARKETING", "Publish ads to TikTok", "TikTok Ads API",
     "TIKTOK_ADS_ACCESS_TOKEN", "Optional", "Push campaigns live to TikTok.",
     "DB-stub only.", "NOT WIRED — publish is DB-stub", ""],
    ["MARKETING", "Publish ads to LinkedIn", "LinkedIn Ads API",
     "LINKEDIN_ADS_ACCESS_TOKEN", "Optional", "Push campaigns live to LinkedIn.",
     "DB-stub only.", "NOT WIRED — publish is DB-stub", ""],
    ["MARKETING", "Publish ads to X", "X Ads API",
     "X_ADS_ACCESS_TOKEN", "Optional", "Push campaigns live to X.",
     "DB-stub only.", "NOT WIRED — publish is DB-stub", ""],

    # ---- BILLING ----
    ["BILLING", "Subscriptions, credit top-ups, webhooks", "Dodo Payments",
     "DODO_PAYMENTS_API_KEY; DODO_WEBHOOK_SECRET; DODO_ENVIRONMENT; DODO_PRODUCT_*_ID; DODO_ADDON_*_ID",
     "REQUIRED to charge", "Real checkout, plan management, credit grants.",
     "Billing/checkout disabled.", "YES", ""],

    # ---- INFRA ----
    ["INFRA", "Error tracking", "Sentry",
     "SENTRY_DSN; NEXT_PUBLIC_SENTRY_DSN; SENTRY_TRACES_SAMPLE_RATE", "Optional",
     "Production error monitoring.", "Inert (no tracking).", "YES", ""],
    ["INFRA", "Protect /api/cron/* endpoints", "Shared secret",
     "CRON_SECRET", "REQUIRED in prod", "Locks cron routes to your scheduler.",
     "Cron endpoints unprotected.", "YES", ""],
    ["INFRA", "App URL", "—",
     "NEXT_PUBLIC_APP_URL", "REQUIRED in prod", "Correct OAuth redirects + links.",
     "Defaults to localhost.", "n/a", ""],
]

for r in rows:
    ws.append(r)
    row_idx = ws.max_row
    area_fill = fill_by_area.get(r[0], INFRA)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.alignment = wrap; cell.border = border
        if c == 1:
            cell.fill = area_fill; cell.font = Font(bold=True)
        if c == 9:
            cell.fill = PatternFill("solid", fgColor="FEF9C3")  # highlight the fill-in column

widths = [15, 34, 26, 34, 16, 40, 40, 26, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# ======================================================================
# Sheet 2 — Verdict / Read me
# ======================================================================
v = wb.create_sheet("Verdict (read first)")
v.column_dimensions["A"].width = 30
v.column_dimensions["B"].width = 95

def hrow(txt):
    v.append([txt, ""])
    cell = v.cell(row=v.max_row, column=1)
    cell.fill = HDR; cell.font = white
    v.merge_cells(start_row=v.max_row, start_column=1, end_row=v.max_row, end_column=2)

def row(a, b):
    v.append([a, b])
    v.cell(row=v.max_row, column=1).font = Font(bold=True)
    v.cell(row=v.max_row, column=1).alignment = wrap
    v.cell(row=v.max_row, column=2).alignment = wrap

hrow("VELDO — Live check summary (as of this run)")
row("1. Is the UI built?",
    "YES. Full Next.js app. Every priority page renders populated (Dashboard, Fundraising, "
    "Marketing, Leads, Campaigns, CRM, Settings all return HTTP 200 with real feature UI). "
    "It is NOT void — it shows clearly-fake demo data (Acme Growth Co / Northwind / Globex).")
row("2. Why demo data?",
    "'.env.local' forces VELDO_FORCE_DEMO=1, so the whole UI is browsable with sample data and no DB. "
    "This is intentional for local preview.")
row("3. Will API keys give real results?",
    "MOSTLY YES — but two gates first (below). The code makes REAL HTTP calls to Anthropic, OpenAI, "
    "Apollo, Tavily, ZeroBounce, Vapi/Bland/Retell, Dodo, and Google. Add the key -> real output.")
row("GATE A — leave demo mode",
    "Remove VELDO_FORCE_DEMO=1 from .env.local AND set the 3 Supabase keys. Otherwise keys do nothing "
    "because the app stays on fake data.")
row("GATE B — create the DB tables",
    "Supabase migrations exist in /supabase/migrations but have NOT been applied to a live DB yet. "
    "Run them before real data will save.")

hrow("What works FULLY once keyed (real results expected)")
row("Sales pipeline",
    "Apollo leads -> Tavily research -> AI email writer (hyper-personalization) -> ZeroBounce verify -> "
    "Gmail send + calendar. End-to-end real. This is the strongest pillar.")
row("AI email writer",
    "Real. Uses premium (deep) model for hyper-personalized mode, grounded only in stored research/signals, "
    "charges credits idempotently, persists drafts. Confirmed in code.")
row("Fundraising (AI parts)",
    "Pitch drafting, outreach, readiness scoring are real AI. Investor MATCHING works on a curated list.")
row("Billing", "Dodo Payments checkout/credits/webhooks are real HTTP.")
row("Voice calling", "Real via Vapi/Bland/Retell when VOICE_PROVIDER_API_KEY is set.")

hrow("What is NOT wired yet (keys won't make these real without code work)")
row("Ad creative media (fal.ai)",
    "Ad COPY is real AI. But the fal.ai image/video job is not actually called — media URL always returns empty. "
    "Needs the fal job implemented.")
row("Ad publishing (Meta/Google/TikTok/LinkedIn/X)",
    "Publishing only writes a row and marks it 'published' in the DB via cron. No real call to any ad platform yet.")
row("Live investor database",
    "INVESTOR_DB_API_KEY is read as a label only; sourcing always returns the built-in curated list. "
    "No external investor API call yet.")
row("E-sign", "Provider interface exists; treat as partial until tested against a real e-sign vendor.")

hrow("Bottom line")
row("Priority = Sales, then Fundraising",
    "Sales pillar is production-ready pending Gate A + Gate B + keys (Apollo, Tavily, ZeroBounce, Google, Anthropic). "
    "Fundraising works except live investor sourcing. Marketing GENERATES but does not EXECUTE (publish) yet.")

for r_ in range(1, v.max_row + 1):
    v.row_dimensions[r_].height = 42

out = r"C:\Users\shind\Desktop\Veldo\Veldo-API-Requirements.xlsx"
wb.save(out)
print("SAVED:", out)
